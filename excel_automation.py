import pandas as pd
import numpy as np

# opening the work file

df = pd.read_csv(r'C:\Users\DELL\Desktop\Python\project\sales_data_sample.csv', encoding='Latin-1')


# to update the object to datetime and formating into date

df['ORDERDATE'] = pd.to_datetime(df['ORDERDATE']).dt.strftime('%d/%m/%Y')

# validation the phone number and convert object into integer

def normalize_phone(phone):
    if phone.startswith('+'):
        phone = phone.split(' ', 1)[-1]
    elif phone.startswith('('):
        phone = phone.split(')', 1)[-1]
    
    phone = phone.replace('.', '').replace(' ', '').replace('-', '')
    return phone

df['PHONE'] = df['PHONE'].apply(normalize_phone).astype(str).astype('int64')

# changing the values 

df['STATUS'] = np.where(df['STATUS'] == 'On Hold', 'In Process', np.where(df['STATUS'] == 'Resolved', 'Disputed', df['STATUS']))

# deleting column of ADDRESSLINE2

df = df.drop(columns=['ADDRESSLINE2'])

# creating a new column

df['CONTACTNAME'] = df['CONTACTFIRSTNAME'] + ' ' + df['CONTACTLASTNAME']

# arranging of columns

cols = list(df.columns)

df = df[cols[:cols.index('CONTACTLASTNAME')] + [cols[-1]] + cols[cols.index('CONTACTLASTNAME'):cols.index(cols[-2])+1]]

# filtering data

df['POSTALCODE'] = df['POSTALCODE'].apply(lambda x: int(x) if isinstance(x, str) and len(x) > 4 and x.isnumeric() else np.nan)

# filling empty cell to value

df['STATE']= df['STATE'].fillna('Others')

import os
if os.path.exists('sales_data_sample.xlsx'):
    os.remove('sales_data_sample.xlsx')

# creating new excel file

df.to_excel('sales_data_sample.xlsx' , index=False)
    
import openpyxl

from openpyxl.styles import PatternFill
 
# opening of excel work file 

wb = openpyxl.load_workbook(r"C:\Users\DELL\Desktop\Python\sales_data_sample.xlsx")

# opening of sheet in excel file

ws = wb['Sheet1']

# finding of index of status column

index_status = [i.value for i in list(ws.iter_rows(max_row=1))[0]].index('STATUS')

color_mapping = {'In Process': 'fcfaa7', 'Disputed': 'b8ffcb', 'Shipped': 'bbd1fa', 'Cancelled': 'fcb3b6'
}

# Iterate through rows, starting from the second row and adding color to row based on values in status column

for row in ws.iter_rows(min_row=2):
    status = row[index_status].value
    if status in color_mapping:
        fill = PatternFill(patternType='solid', fgColor=color_mapping[status])
        for cell in row:
            cell.fill = fill

import os
if os.path.exists('sales_data_update.xlsx'):
    os.remove('sales_data_update.xlsx')

# creating new excel file

wb.save("sales_data_update.xlsx" )